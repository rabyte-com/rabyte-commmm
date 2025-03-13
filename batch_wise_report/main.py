from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from datetime import datetime, timedelta
from hdbcli import dbapi
from openpyxl.styles import Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
import smtplib
import pandas as pd
import os

current_date = (datetime.now()).strftime("%d.%m.%Y")


def get_data():
    # connection to PTE server
    cnx1 = dbapi.connect(
        address="103.25.172.160",
        port=30015,
        database="RABYTE_PTE_LIVE",
        user="SYSTEM",
        password="Data1234",
    )
    print("🟢 PTE Connection Done")
    query1 = """
        SELECT DISTINCT 
            T0."ItemCode", 
            T0."ItemName", 
            T6."Name" as "Make",
            T11."Name" "Category-2", 
            T0."WhsCode", 
            T0."BatchNum", 
            CASE 
                WHEN T2."DocNum" IS NULL THEN CURRENT_TIMESTAMP
                ELSE T3."CreateDate"
            END as "Allocation Ageing Days",
            T3."InDate" as "Inventory Ageing Days", 
            T2."DocNum" as "SO Number", 
            T1."U_CRDDate" as "CRD", 
            T2."CardCode" as "Cust. Code",
            T2."CardName" as "Cust. Name", 
            T4."Quantity", 
            T5."LastPurPrc" as "Mov. Avg. Price",
            (T4."Quantity" * T5."LastPurPrc") as "Total Amount"
        FROM 
            "RABYTE_PTE_LIVE".IBT1 T0
        LEFT JOIN 
            "RABYTE_PTE_LIVE".OBTN T3 ON T0."ItemCode" = T3."ItemCode" AND T0."BatchNum" = T3."DistNumber"
        LEFT JOIN 
            "RABYTE_PTE_LIVE".OIBT T4 ON T0."WhsCode" = T4."WhsCode" AND T0."ItemCode" = T4."ItemCode" AND T0."BatchNum" = T4."BatchNum"
        LEFT JOIN 
            "RABYTE_PTE_LIVE".RDR1 T1 ON 'SO_' || T1."DocEntry" || '_' || T1."LineNum" = T0."BatchNum" 
            OR 'SO_' || T1."DocEntry" || '_' || T1."LineNum" = T3."LotNumber" 
        LEFT JOIN 
            "RABYTE_PTE_LIVE".ORDR T2 ON T1."DocEntry" = T2."DocEntry"
        LEFT JOIN 
            "RABYTE_PTE_LIVE".OITM T5 ON T5."ItemCode" = T0."ItemCode"
        LEFT JOIN 
            "RABYTE_PTE_LIVE"."@MAKE_FULL_2" T6 ON T6."Code" = T5."U_M_F_2"
        LEFT JOIN 
            "RABYTE_PTE_LIVE"."@CATEGORY_2" T11 ON T5.U_CAT2 = T11."Code"
        WHERE 
            T4."Quantity" > 0
    """

    cursor1 = cnx1.cursor()
    cursor1.execute(query1)
    data = cursor1.fetchall()
    column_names1 = [i[0] for i in cursor1.description]
    df1 = pd.DataFrame(data, columns=column_names1)
    df1["Allocation Ageing Days"] = pd.to_datetime(df1["Allocation Ageing Days"])
    df1["Allocation Ageing Days"] = (
        pd.to_datetime("today") - df1["Allocation Ageing Days"]
    ).dt.days
    df1["Inventory Ageing Days"] = pd.to_datetime(df1["Inventory Ageing Days"])
    df1["Inventory Ageing Days"] = (
        pd.to_datetime("today") - df1["Inventory Ageing Days"]
    ).dt.days
    df1["CRD"] = pd.to_datetime(df1["CRD"]).dt.date
    df1["Quantity"] = df1["Quantity"].astype(int)
    df1["Mov. Avg. Price"] = df1["Mov. Avg. Price"].astype(float)
    df1["Total Amount"] = df1["Total Amount"].astype(float)
    df1["Company"] = "PTE"
    df1 = df1[["Company"] + [col for col in df1.columns if col != "Company"]]

    # connection to LLP server
    cnx2 = dbapi.connect(
        address="103.25.172.160",
        port=30015,
        database="RABYTE_LLP_LIVE01_0808",
        user="SYSTEM",
        password="Data1234",
    )
    print("🟢 LLP Connection Done")
    query2 = """
        SELECT DISTINCT 
            T0."ItemCode", 
            T0."ItemName", 
            T6."Name" as "Make",
            T11."Name" "Category-2", 
            T0."WhsCode", 
            T0."BatchNum",
            CASE 
                WHEN T2."DocNum" IS NULL THEN CURRENT_TIMESTAMP
                ELSE T3."CreateDate"
            END as "Allocation Ageing Days",
            T3."InDate" as "Inventory Ageing Days", 
            T2."DocNum" as "SO Number", 
            T1."U_CRDDate" as "CRD", 
            T2."CardCode" as "Cust. Code",
            T2."CardName" as "Cust. Name", 
            T4."Quantity", 
            T5."LastPurPrc" as "Mov. Avg. Price",
            (T4."Quantity" * T5."LastPurPrc")/TT."Rate" as "Total Amount"
        FROM 
            "RABYTE_LLP_LIVE01_0808".IBT1 T0
        LEFT JOIN 
            "RABYTE_LLP_LIVE01_0808".OBTN T3 ON T0."ItemCode" = T3."ItemCode" AND T0."BatchNum" = T3."DistNumber"
        LEFT JOIN 
            "RABYTE_LLP_LIVE01_0808".OIBT T4 ON T0."WhsCode" = T4."WhsCode" AND T0."ItemCode" = T4."ItemCode" AND T0."BatchNum" = T4."BatchNum"
        LEFT JOIN 
            "RABYTE_LLP_LIVE01_0808".RDR1 T1 ON 'SO_' || T1."DocEntry" || '_' || T1."LineNum" = T0."BatchNum" 
            OR 'SO_' || T1."DocEntry" || '_' || T1."LineNum" = T3."LotNumber" 
        LEFT JOIN 
            "RABYTE_LLP_LIVE01_0808".ORDR T2 ON T1."DocEntry" = T2."DocEntry"
        LEFT JOIN 
            "RABYTE_LLP_LIVE01_0808".OITM T5 ON T5."ItemCode" = T0."ItemCode"
        LEFT JOIN 
            "RABYTE_LLP_LIVE01_0808"."@MAKE_FULL_2" T6 ON T6."Code" = T5."U_M_F_2"
        LEFT JOIN 
            "RABYTE_LLP_LIVE01_0808"."@CATEGORY_2" T11 ON T5.U_CAT2 = T11."Code"
        LEFT JOIN 
            "RABYTE_LLP_LIVE01_0808".ORTT TT ON TT."RateDate" = CURRENT_DATE AND TT."Currency" = 'USD'
        WHERE 
            T4."Quantity" > 0
    """

    cursor2 = cnx2.cursor()
    cursor2.execute(query2)
    data = cursor2.fetchall()
    column_names2 = [i[0] for i in cursor2.description]
    df2 = pd.DataFrame(data, columns=column_names2)
    df2["Allocation Ageing Days"] = pd.to_datetime(df2["Allocation Ageing Days"])
    df2["Allocation Ageing Days"] = (
        (pd.to_datetime("today") - df2["Allocation Ageing Days"])
    ).dt.days
    df2["Inventory Ageing Days"] = pd.to_datetime(df2["Inventory Ageing Days"])
    df2["Inventory Ageing Days"] = (
        pd.to_datetime("today") - df2["Inventory Ageing Days"]
    ).dt.days
    df2["CRD"] = pd.to_datetime(df2["CRD"]).dt.date
    df2["Total Amount"] = df2["Total Amount"].astype("float64").round(3)
    df2["Quantity"] = df2["Quantity"].astype(int)
    df2["Mov. Avg. Price"] = df2["Mov. Avg. Price"].astype(float)
    df2["Total Amount"] = df2["Total Amount"].astype(float)
    df2["Company"] = "LLP"
    df2 = df2[["Company"] + [col for col in df2.columns if col != "Company"]]

    # concatinating and saving in excel file
    excel = pd.concat([df2, df1], ignore_index=True)
    excel.to_excel(
        f"files/output {current_date}.xlsx", index=False, sheet_name="Batch_Wise_Report"
    )
    wb = load_workbook(f"files/output {current_date}.xlsx")
    ws = wb.active
    border_style = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    header_fill = PatternFill(
        start_color="1ec71e", end_color="1ec71e", fill_type="solid"
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.border = border_style
        cell.alignment = Alignment(horizontal="left", vertical="bottom")
        column_letter = cell.column_letter
        ws.column_dimensions[column_letter].width = 15
    wb.save(f"files/output {current_date}.xlsx")
    print("🟡File Saved")


def send_email(body, attachment_path=None):
    """Send an email with optional attachment."""
    smtp_server = "smtp.office365.com"
    smtp_port = 587
    smtp_user = "crm@rabyte.com"
    smtp_password = "rPB7D7#5rP"
    from_email = "crm@rabyte.com"
    to_email = ["harimohan.pathak.ug@rabyte.com"]
    cc_email = ["mudit.gupta.uv@rabyte.com"]
    filename = f"Batch_Wise_Stock_Report {current_date}"
    msg = MIMEMultipart()
    msg["To"] = ", ".join(to_email)
    msg["Cc"] = ", ".join(cc_email)
    msg["subject"] = f"Batch Wise Stock Report on dated - {current_date}"
    # Attach body
    msg.attach(MIMEText(body, "html"))

    # Attach file if present
    if attachment_path:
        with open(attachment_path, "rb") as file:
            part = MIMEApplication(file.read(), Name="output.xlsx")
            part["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
            msg.attach(part)

    # Send the email
    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(smtp_user, smtp_password)
        rpients = to_email + cc_email
        server.sendmail(from_email, rpients, msg.as_string())

    print("🟡🟢 Mail Sent")


if __name__ == "__main__":
    print("✅✅Starting")
    get_data()
    html = """
        <html>
            <body>
                <p style="font-size:15px">Dear Sir,<br><br>Please find the attached batch wise stock report of both company(LLP and PTE).</p>
                <p style="font-style:italic;font-size:9px;margin: 0px 0px 0px 0px;"><strong style="font-size:9px;">Data Source</strong> – Direct from SAP (Both Company).</p>
                <p><strong>Regards,</strong><br>
                <br>
                <font color="black"><small><i><strong>Note -</strong> This is an auto-generated email. If you require any technical support, please contact the IT-CRM Team.</i></small></font></p>
            </body>
        </html>
    """
    # send_email(html, f"files/output {current_date}.xlsx")
    # os.remove("files/output.xlsx")
    print("🔴🔴Ending")
