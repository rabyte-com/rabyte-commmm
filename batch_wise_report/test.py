from datetime import datetime
from hdbcli import dbapi
from openpyxl.styles import Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
import pandas as pd

current_date = (datetime.now()).strftime("%d.%m.%Y_%H%M%S")

custom_columns = [
    "ItemCode",
    "ItemName",
    "Make",
    "Make+Div",
    "Planning Person",
    "Cat2",
    "Cat3",
    "WhsCode",
    "BatchNum",
    "Allocation Ageing Days",
    "Inventory Ageing Days",
    "SO Number",
    "Sales Head",
    "Sales Person",
    "CRD",
    "Cust. Code",
    "Cust. Name",
    "Quantity",
    "Mov. Avg. Price",
    "Sales Price",
    "Total Amount ($)",
]


def get_data():
    # connection to PTE server
    cnx1 = dbapi.connect(
        address="103.25.172.160",
        port=30015,
        database="RABYTE_PTE_LIVE",
        user="SYSTEM",
        password="Data1234",
    )
    print("🟢 PTE Connection Done")
    query1 = """
        SELECT DISTINCT 
            T0."ItemCode",
            T0."ItemName",
            T7."Name" "Make",
            T6."Name" as "Make+Div",
            T11."Name" AS "Cat2", 
            T12."Name" as "Cat3",
            T0."WhsCode",
            T0."BatchNum",
            CASE 
                WHEN T2."DocNum" IS NULL THEN CURRENT_TIMESTAMP
                ELSE T3."CreateDate"
            END as "Allocation Ageing Days",
            T3."InDate" as "Inventory Ageing Days",
            T2."DocNum" as "SO Number",
            T1."U_CRDDate" as CRD,
            T2."CardCode" as "Cust. Code",
            T2."CardName" as "Cust. Name", 
            T4."Quantity",
            T5."LastPurPrc" as "Mov. Avg. Price",
            (T4."Quantity" * T5."LastPurPrc") as "Total Amount ($)"

            FROM 
            "RABYTE_PTE_LIVE".IBT1  T0
            
            LEFT JOIN 
            "RABYTE_PTE_LIVE".OBTN T3 ON T0."ItemCode"=T3."ItemCode" AND T0."BatchNum"=T3."DistNumber"

            LEFT JOIN 
            "RABYTE_PTE_LIVE".OIBT T4 ON T0."WhsCode" = T4."WhsCode"and T0."ItemCode" = T4."ItemCode" AND T0."BatchNum"=T4."BatchNum"

            LEFT JOIN 
            "RABYTE_PTE_LIVE".RDR1 T1 ON 'SO_' || T1."DocEntry"||'_'||T1."LineNum"=T0."BatchNum" OR   'SO_' || T1."DocEntry"||'_'||T1."LineNum"=T3."LotNumber"   

            LEFT JOIN 
            "RABYTE_PTE_LIVE".ORDR T2 ON T1."DocEntry"=T2."DocEntry"

            LEFT JOIN 
            "RABYTE_PTE_LIVE".OITM T5 ON T5."ItemCode"=T0."ItemCode"

            LEFT JOIN 
            "RABYTE_PTE_LIVE"."@MAKE_FULL_2" T6 ON T6."Code"= T5."U_M_F_2"

            LEFT JOIN 
            "RABYTE_PTE_LIVE"."@MAKE_FULL_1" T7 ON T7."Code"= T5."U_MF_1"

            LEFT JOIN 
            "RABYTE_PTE_LIVE"."@CATEGORY_2" T11 ON T5."U_CAT2"=T11."Code"  

            LEFT JOIN 
            "RABYTE_PTE_LIVE"."@CATEGORY_3" T12 ON T5."U_CAT3"=T12."Code"  

            WHERE T4."Quantity" >0
    """

    cursor1 = cnx1.cursor()
    cursor1.execute(query1)
    data = cursor1.fetchall()
    column_names1 = [i[0] for i in cursor1.description]
    df1 = pd.DataFrame(data, columns=column_names1)
    df1["Allocation Ageing Days"] = pd.to_datetime(df1["Allocation Ageing Days"])
    df1["Allocation Ageing Days"] = (
        pd.to_datetime("today") - df1["Allocation Ageing Days"]
    ).dt.days
    df1["Inventory Ageing Days"] = pd.to_datetime(df1["Inventory Ageing Days"])
    df1["Inventory Ageing Days"] = (
        pd.to_datetime("today") - df1["Inventory Ageing Days"]
    ).dt.days
    df1["CRD"] = pd.to_datetime(df1["CRD"]).dt.date
    df1["Quantity"] = df1["Quantity"].astype(int)
    df1["Mov. Avg. Price"] = df1["Mov. Avg. Price"].astype(float)
    df1["Total Amount ($)"] = df1["Total Amount ($)"].astype(float)
    df1["Sales Price"] = ""
    df1["Planning Person"] = ""
    df1["Sales Head"] = ""
    df1["Sales Person"] = ""
    df1["Company"] = "PTE"
    df1 = df1[["Company"] + [col for col in df1.columns if col != "Company"]]
    df1 = df1[custom_columns]
    
    # connection to LLP server
    cnx2 = dbapi.connect(
        address="103.25.172.160",
        port=30015,
        database="RABYTE_LLP_LIVE01_0808",
        user="SYSTEM",
        password="Data1234",
    )
    print("🟢 LLP Connection Done")
    query2 = """
        SELECT DISTINCT 
            T0."ItemCode",
            T0."ItemName",
            T7."Name" "Make",
            T6."Name" as "Make+Div",
            T11."Name" AS "Cat2", 
            T12."Name" as "Cat3",
            T0."WhsCode",
            T0."BatchNum",
            CASE 
                WHEN T2."DocNum" IS NULL THEN CURRENT_TIMESTAMP
                ELSE T3."CreateDate"
            END as "Allocation Ageing Days",
            T3."InDate" as "Inventory Ageing Days",
            T2."DocNum" as "SO Number",
            T1."U_CRDDate" as CRD,
            T2."CardCode" as "Cust. Code",
            T2."CardName" as "Cust. Name", 
            T4."Quantity",
            T5."LastPurPrc" as "Mov. Avg. Price",
            (T4."Quantity" * T5."LastPurPrc")/TT."Rate" as "Total Amount ($)"

            FROM 
            "RABYTE_LLP_LIVE01_0808".IBT1  T0
            LEFT JOIN 
            "RABYTE_LLP_LIVE01_0808".OBTN T3 ON T0."ItemCode"=T3."ItemCode" AND T0."BatchNum"=T3."DistNumber"
            LEFT JOIN 
            "RABYTE_LLP_LIVE01_0808".OIBT T4 ON T0."WhsCode" = T4."WhsCode"and T0."ItemCode" = T4."ItemCode" AND T0."BatchNum"=T4."BatchNum"
            LEFT JOIN 
            "RABYTE_LLP_LIVE01_0808".RDR1 T1 ON 'SO_' || T1."DocEntry"||'_'||T1."LineNum"=T0."BatchNum" OR   'SO_' || T1."DocEntry"||'_'||T1."LineNum"=T3."LotNumber"   
            LEFT JOIN 
            "RABYTE_LLP_LIVE01_0808".ORDR T2 ON T1."DocEntry"=T2."DocEntry"

            LEFT JOIN 
            "RABYTE_LLP_LIVE01_0808".OITM T5 ON T5."ItemCode"=T0."ItemCode"

            LEFT JOIN 
            "RABYTE_LLP_LIVE01_0808"."@MAKE_FULL_2" T6 ON T6."Code"= T5."U_M_F_2"

            LEFT JOIN 
            "RABYTE_LLP_LIVE01_0808"."@MAKE_FULL_1" T7 ON T7."Code"= T5."U_MF_1"

            LEFT JOIN 
            "RABYTE_LLP_LIVE01_0808"."@CATEGORY_2" T11 ON T5."U_CAT2"=T11."Code"  

            LEFT JOIN 
            "RABYTE_LLP_LIVE01_0808"."@CATEGORY_3" T12 ON T5."U_CAT3"=T12."Code"  

            LEFT JOIN 
            "RABYTE_LLP_LIVE01_0808".ORTT TT ON TT."RateDate" = CURRENT_DATE AND TT."Currency" = 'USD'

            WHERE T4."Quantity" >0
    """

    cursor2 = cnx2.cursor()
    cursor2.execute(query2)
    data = cursor2.fetchall()
    column_names2 = [i[0] for i in cursor2.description]
    df2 = pd.DataFrame(data, columns=column_names2)
    df2["Allocation Ageing Days"] = pd.to_datetime(df2["Allocation Ageing Days"])
    df2["Allocation Ageing Days"] = (
        (pd.to_datetime("today") - df2["Allocation Ageing Days"])
    ).dt.days
    df2["Inventory Ageing Days"] = pd.to_datetime(df2["Inventory Ageing Days"])
    df2["Inventory Ageing Days"] = (
        pd.to_datetime("today") - df2["Inventory Ageing Days"]
    ).dt.days
    df2["CRD"] = pd.to_datetime(df2["CRD"]).dt.date
    df2["Total Amount ($)"] = df2["Total Amount ($)"].astype("float64").round(3)
    df2["Quantity"] = df2["Quantity"].astype(int)
    df2["Mov. Avg. Price"] = df2["Mov. Avg. Price"].astype(float)
    df2["Total Amount ($)"] = df2["Total Amount ($)"].astype(float)
    df2["Sales Price"] = ""
    df2["Planning Person"] = ""
    df2["Sales Head"] = ""
    df2["Sales Person"] = ""
    df2["Company"] = "LLP"
    df2 = df2[["Company"] + [col for col in df2.columns if col != "Company"]]
    df2 = df2[custom_columns]

    # concatinating and saving in excel file
    excel = pd.concat([df2, df1], ignore_index=True)
    excel.to_excel(
        f"files/output {current_date}.xlsx", index=False, sheet_name="Batch_Wise_Report"
    )
    wb = load_workbook(f"files/output {current_date}.xlsx")
    ws = wb.active
    border_style = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    header_fill = PatternFill(
        start_color="1ec71e", end_color="1ec71e", fill_type="solid"
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.border = border_style
        cell.alignment = Alignment(horizontal="left", vertical="bottom")
        column_letter = cell.column_letter
        ws.column_dimensions[column_letter].width = 18
    wb.save(f"files/output {current_date}.xlsx")
    print("🟡File Saved")


if __name__ == "__main__":
    print("✅✅Starting")
    get_data()
    print("🔴🔴Ending")
