from datetime import datetime
from hdbcli import dbapi
from openpyxl.styles import Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
import pandas as pd

current_date = (datetime.now()).strftime("%d.%m.%Y")


def get_data():
    # connection to LLP server
    cnx2 = dbapi.connect(
        address="103.25.172.160",
        port=30015,
        database="RABYTE_LLP_LIVE01_0808",
        user="SYSTEM",
        password="Data1234",
    )
    print("🟢 LLP Connection Done")
    query = f"""
        SELECT DISTINCT T1."DocNum" AS "Purchase order number",T1."DocDate" as "Purchase order date",T1."NumAtCard" "Contract/Quote number/Quote Line Item",

        '' "Buyer/End Customer code/Ship To",'' "Assigned by seller/Buyer",'' "Party Name - STREET/City/Region/postalcode/country",

        (SELECT PY."PymntGroup" FROM "RABYTE_PTE_LIVE".OCTG PY WHERE PY."GroupNum" = T1."GroupNum")"Payment Terms",T1."U_Price" AS "Delivery Term",

        '' "LINE ITEM No",T0."ItemCode" as "Supplier’s Part Number",T0."ItemCode" as "Vendor's part number",'' as "Assigned by seller/Buyer",

        T0."Dscription"  as "Customer Part No",T0."ItemCode" as "Distributor’s Part Number",

        T0."Quantity" AS "Ordered quantity",T0."Quantity" AS "Ordered total quantity",T0."PriceBefDi" AS "Distributor Unit price",

        '' as "Buyer's original line item Number",'' as "Contract/Quote number",'' as "Buyer's original line item number",

        '' as "SCHEDULING CONDITIONS",'' as "Ordered quantity",'' as "Schedule QTY",

        T0."U_CRDDate" as "Buyer Delivery date/time",'' AS "Contract/Quote Ref if applicable at HEADER level",'' as "Buyer Code",

        T1."U_OSRAMEndCustNo" as "Disty Code given by supplier (End Customer Code)",T41."Currency",

        T0."ItemCode" as "Vendor Part No",''  as "Customer Part No",'PCS' as "Price And Unit Basic",

        '' as "Customer Line Item No",T1."DocStatus" "PO Status",T11."DocStatus" as "So Status",T0."Dscription",T1."DocCur",T0."U_OsramQuoteNo",T0."U_OsramLineNo"

        FROM "RABYTE_PTE_LIVE".POR1 T0  

        INNER JOIN "RABYTE_PTE_LIVE".OPOR T1 ON T0."DocEntry" = T1."DocEntry" AND T1."CANCELED"='N'

        INNER  JOIN "RABYTE_PTE_LIVE".OITM T4 ON T0."ItemCode"=T4."ItemCode"

        LEFT JOIN "RABYTE_PTE_LIVE".OITB T24 ON T24."ItmsGrpCod"=T4."ItmsGrpCod"

        LEFT OUTER JOIN "RABYTE_PTE_LIVE".PDN1 T2 ON T0."DocEntry" = T2."BaseEntry" AND T0."LineNum" = T2."BaseLine" AND T2."ItemCode" = T0."ItemCode"

        LEFT OUTER JOIN "RABYTE_PTE_LIVE".OPDN T3 ON T2."DocEntry"=T3."DocEntry" AND T3."CANCELED"='N'

        LEFT OUTER JOIN "RABYTE_PTE_LIVE"."@MAKE_FULL_1" T5 ON T4."U_MF_1"= T5."Code"

        LEFT OUTER JOIN "RABYTE_PTE_LIVE"."@MAKE_FULL_2" T15 ON T4."U_M_F_2"=T15."Code"

        LEFT OUTER JOIN "RABYTE_PTE_LIVE"."@UTL_PODHED" T10 ON T0."DocEntry"=T10."U_BASEENTRY" AND T0."LineNum"=T10."U_BASELINE"

        LEFT OUTER JOIN "RABYTE_PTE_LIVE"."@UTL_PODDET" T9 ON T10."DocEntry"=T9."DocEntry" AND T9."U_SOQTY">0

        LEFT OUTER JOIN "RABYTE_PTE_LIVE".OITW T8 ON T4."ItemCode"=T8."ItemCode" AND T8."ItemCode"=T0."ItemCode"

        LEFT OUTER JOIN "RABYTE_PTE_LIVE"."@UTL_SODHED" T6 ON T9."U_SODOCENTRY"=T6."U_BASEENTRY" AND T6."U_BASELINE"=T9."U_SOLINE"

        LEFT OUTER JOIN "RABYTE_PTE_LIVE"."@UTL_SODDET" T7 ON T6."DocEntry"=T7."DocEntry"

        LEFT OUTER JOIN "RABYTE_PTE_LIVE".RDR1 T12 ON T9."U_SOLINE"=T12."LineNum" AND T9."U_SODOCENTRY"=T12."DocEntry"

        LEFT OUTER JOIN "RABYTE_PTE_LIVE".ORDR T11 ON T12."DocEntry"=T11."DocEntry"

        LEFT OUTER JOIN "RABYTE_PTE_LIVE"."@CATEGORY_2" T13 ON T4."U_CAT2"=T13."Code"

        LEFT OUTER JOIN "RABYTE_PTE_LIVE"."@CATEGORY_3" T14 ON T4."U_CAT3"=T14."Code"

        LEFT OUTER JOIN "RABYTE_PTE_LIVE".OSLP T16 ON T11."SlpCode"=T16."SlpCode"

        LEFT JOIN "RABYTE_PTE_LIVE".OCRD T23 ON T23."CardCode"=T11."CardCode"

        LEFT OUTER JOIN "RABYTE_PTE_LIVE".OCHP T17 ON T4."ChapterID"=T17."AbsEntry"

        LEFT OUTER JOIN "RABYTE_PTE_LIVE"."@CATEGORY_4" T20 ON T4."U_CAT4"=T20."Code"

        LEFT OUTER JOIN "RABYTE_PTE_LIVE"."@CATEGORY_5" T21 ON T4."U_CAT5"=T21."Code"

        LEFT OUTER JOIN "RABYTE_PTE_LIVE".QUT1 T19 ON T12."BaseEntry"=T19."DocEntry" AND T12."BaseLine"=T19."LineNum"

        LEFT OUTER JOIN "RABYTE_PTE_LIVE".OQUT T22 ON T22."DocEntry"= T19."DocEntry"

        LEFT OUTER JOIN "RABYTE_PTE_LIVE".NNM1 T40 ON T40."Series"=T1."Series"

        LEFT OUTER JOIN "RABYTE_PTE_LIVE".OCRD T41 ON T41."CardCode"=T11."CardCode"

        where T1."U_MGMTDate"='{current_date}' and T5."Name"='ams Osram' and T0."LineStatus"='O' and T1."U_MGMTAPPROVAL"='Yes'

        GROUP BY T1."DocNum", T1."DocDate", T1."CardCode",T1."Comments",T0."Dscription",T0."ItemCode",T0."Dscription",Cast(T0."Text" as Varchar(20)),T5."Name",

        T1."DocEntry", T0."Quantity",T0."FreeTxt",T0."U_Remarks2",T0."U_Remarks3",T0."U_Remarks4",T17."ChapterID",T2."Quantity",T0."U_CRDDate",

        T4."LeadTime",T0."PriceBefDi" ,T0."LineTotal",T4."OnHand",T11."DocNum",T0."U_UTL_PMAprvl",T0."U_UTL_OthAprvl",T1."U_UTL_Status",

        T0."LineNum",T9."U_SOQTY",T9."U_SOLINE",T10."U_BASELINE",T12."LineNum",T12."Quantity",T12."ShipDate",T11."DocNum",T0."WhsCode",

        T1."CardName",T0."U_UTL_BPREF",T1."NumAtCard",T13."Name",T14."Name" ,T15."U_PMName",T15."U_PMAName",T0."U_InStock",T15."Name",

        T11."CardCode",T11."CardName",T11."NumAtCard",T12."U_UTL_BPREF",T12."U_CRDDate",T16."SlpName",T17."ChapterID",T0."FreeTxt",T0."LineStatus",T9."U_AQTY",

        T0."U_Remarks2",T0."U_Remarks3",T0."U_Remarks4" ,T20."Name",T21."Name",T19."LineNum",T22."DocNum",T12."U_UTL_AQTY",T12."U_SuppEDD",T1."DocStatus",T0."LineStatus",T40."SeriesName",

        T0."U_FOC" ,T0."U_FOCTyp" ,T0."U_SAMPLETYPE" ,T23."U_SALES_MANAGER",

        T24."ItmsGrpNam" ,T1."U_OrderNo",T1."DocCur",

        T0."U_UTL_RENESCMNO",T0."U_SuppEDD",T0."OpenCreQty",T11."Address2",T1."GroupNum",T1."U_Price"

        ,T41."Currency",T41."U_OsramCustNo",T1."DocStatus",T11."DocStatus",T12."U_UTL_CPN",T1."U_OSRAMEndCustNo",T1."NumAtCard",T0."Dscription",T1."DocCur",T0."U_OsramQuoteNo",T0."U_OsramLineNo"
        LIMIT 5
        """

    cursor2 = cnx2.cursor()
    cursor2.execute(query)
    data = cursor2.fetchall()
    column_names2 = [i[0] for i in cursor2.description]
    df2 = pd.DataFrame(data, columns=column_names2)

    # concatinating and saving in excel file
    excel = pd.concat([df2, df2], ignore_index=True)
    excel.to_excel(
        f"files/output po {current_date}.xlsx", index=False, sheet_name="Batch_Wise_Report"
    )
    wb = load_workbook(f"files/output po {current_date}.xlsx")
    ws = wb.active
    border_style = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    header_fill = PatternFill(
        start_color="1ec71e", end_color="1ec71e", fill_type="solid"
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.border = border_style
        cell.alignment = Alignment(horizontal="left", vertical="bottom")
        column_letter = cell.column_letter
        ws.column_dimensions[column_letter].width = 15
    wb.save(f"files/output {current_date}.xlsx")
    print("🟡File Saved")

if __name__ == "__main__":
    print("✅✅Starting")
    get_data()
    print("🔴🔴Ending")
